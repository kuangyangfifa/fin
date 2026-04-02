import io
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def export_transactions_to_excel(transactions, account_map=None):
    """Export transactions to Excel file.

    Args:
        transactions: List of Transaction objects
        account_map: Optional dict mapping account_id to account_name

    Returns:
        BytesIO object containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "流水记录"

    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ['ID', '日期', '账户', '摘要', '收入', '支出', '余额', '备注1', '备注2', '备注3', '备注4', '备注5']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    total_income = 0
    total_expense = 0

    for row, trans in enumerate(transactions, 2):
        account_name = account_map.get(trans.account_id, '未知账户') if account_map else str(trans.account_id)

        ws.cell(row=row, column=1, value=trans.id).border = thin_border
        ws.cell(row=row, column=2, value=trans.date.strftime('%Y-%m-%d')).border = thin_border
        ws.cell(row=row, column=3, value=account_name).border = thin_border
        ws.cell(row=row, column=4, value=trans.summary or '').border = thin_border
        ws.cell(row=row, column=5, value=trans.income).border = thin_border
        ws.cell(row=row, column=6, value=trans.expense).border = thin_border
        ws.cell(row=row, column=7, value=trans.balance_after).border = thin_border
        ws.cell(row=row, column=8, value=trans.note1 or '').border = thin_border
        ws.cell(row=row, column=9, value=trans.note2 or '').border = thin_border
        ws.cell(row=row, column=10, value=trans.note3 or '').border = thin_border
        ws.cell(row=row, column=11, value=trans.note4 or '').border = thin_border
        ws.cell(row=row, column=12, value=trans.note5 or '').border = thin_border

        total_income += trans.income
        total_expense += trans.expense

    # Summary row
    summary_row = len(transactions) + 2
    summary_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    summary_font = Font(bold=True)

    ws.cell(row=summary_row, column=3, value="合计:").font = summary_font
    ws.cell(row=summary_row, column=5, value=total_income).fill = summary_fill
    ws.cell(row=summary_row, column=6, value=total_expense).fill = summary_fill
    ws.cell(row=summary_row, column=7, value=total_income - total_expense).fill = summary_fill

    # Adjust column widths
    column_widths = [8, 12, 20, 30, 12, 12, 12, 20, 20, 20, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_import_template():
    """Create Excel template for importing transactions.

    Returns:
        BytesIO object containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"

    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    example_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Headers
    headers = ['日期', '账户名称', '摘要', '收入', '支出', '备注1', '备注2', '备注3', '备注4', '备注5']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Example data (will be highlighted)
    example_data = [
        ['2024-01-15', '现金账户', '销售收入', 1000.00, 0, '', '', '', '', ''],
        ['2024-01-16', '银行账户', '办公用品采购', 0, 200.00, '发票号001', '', '', '', ''],
    ]

    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill

    # Instructions
    ws.cell(row=5, column=1, value="说明：")
    ws.cell(row=5, column=1).font = Font(bold=True)
    ws.cell(row=6, column=1, value="1. 日期格式：YYYY-MM-DD")
    ws.cell(row=7, column=1, value="2. 账户名称必须是系统中已存在的账户")
    ws.cell(row=8, column=1, value="3. 收入和支出只能填写一个，另一个填0")
    ws.cell(row=9, column=1, value="4. 删除示例数据后填写真实数据")

    # Adjust column widths
    column_widths = [12, 20, 30, 12, 12, 20, 20, 20, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def parse_excel_import(file_stream, account_name_to_id):
    """Parse Excel file for importing transactions.

    Args:
        file_stream: File stream of the uploaded Excel file
        account_name_to_id: Dict mapping account names to account IDs

    Returns:
        Tuple of (success: bool, result: list or error message)
    """
    try:
        wb = load_workbook(file_stream)
        ws = wb.active

        transactions = []
        errors = []

        # Skip header row, start from row 2
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            # Skip empty rows
            if not any(row):
                continue

            if len(row) < 5:
                errors.append(f"第{row_idx}行：数据不完整")
                continue

            date_str, account_name, summary, income, expense = row[0:5]
            note1 = row[5] if len(row) > 5 else ''
            note2 = row[6] if len(row) > 6 else ''
            note3 = row[7] if len(row) > 7 else ''
            note4 = row[8] if len(row) > 8 else ''
            note5 = row[9] if len(row) > 9 else ''

            # Validate date
            if not date_str:
                errors.append(f"第{row_idx}行：日期不能为空")
                continue

            try:
                if isinstance(date_str, str):
                    date_obj = datetime.strptime(date_str.strip(), '%Y-%m-%d').date()
                elif isinstance(date_str, datetime):
                    date_obj = date_str.date()
                else:
                    errors.append(f"第{row_idx}行：日期格式错误，请使用 YYYY-MM-DD 格式")
                    continue
            except ValueError:
                errors.append(f"第{row_idx}行：日期格式错误，请使用 YYYY-MM-DD 格式")
                continue

            # Validate account
            if not account_name:
                errors.append(f"第{row_idx}行：账户名称不能为空")
                continue

            account_name = str(account_name).strip()
            if account_name not in account_name_to_id:
                errors.append(f"第{row_idx}行：账户 '{account_name}' 不存在，请先创建该账户")
                continue

            account_id = account_name_to_id[account_name]

            # Validate amounts
            try:
                income = float(income) if income else 0.0
                expense = float(expense) if expense else 0.0
            except (ValueError, TypeError):
                errors.append(f"第{row_idx}行：收入或支出金额格式错误")
                continue

            if income < 0 or expense < 0:
                errors.append(f"第{row_idx}行：收入和支出不能为负数")
                continue

            if income == 0 and expense == 0:
                errors.append(f"第{row_idx}行：收入和支出不能同时为0")
                continue

            transactions.append({
                'date': date_obj,
                'account_id': account_id,
                'summary': str(summary) if summary else '',
                'income': income,
                'expense': expense,
                'note1': str(note1) if note1 else '',
                'note2': str(note2) if note2 else '',
                'note3': str(note3) if note3 else '',
                'note4': str(note4) if note4 else '',
                'note5': str(note5) if note5 else ''
            })

        if errors:
            return False, errors

        if not transactions:
            return False, ["未找到有效的交易数据"]

        # Sort by date
        transactions.sort(key=lambda x: x['date'])

        return True, transactions

    except Exception as e:
        return False, [f"解析Excel文件时出错：{str(e)}"]
